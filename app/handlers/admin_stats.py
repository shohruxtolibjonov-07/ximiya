import io
import csv
from datetime import datetime

from aiogram import Bot, Router, F
from aiogram.types import Message, CallbackQuery, BufferedInputFile
from sqlalchemy.ext.asyncio import AsyncSession

from app.filters.admin import IsAdmin
from app.database import queries as db
from app.keyboards.inline import stats_export_kb
from app.utils import texts

router = Router()
router.message.filter(IsAdmin())
router.callback_query.filter(IsAdmin())


@router.message(F.text == "📊 Statistika")
async def show_stats(message: Message, session: AsyncSession):
    """Statistikani ko'rsatish."""
    total = await db.get_user_count(session)
    akademiya = await db.get_user_count_by_center(session, "Akademiya")
    ilmnur = await db.get_user_count_by_center(session, "Ilm-nur")

    all_resources = await db.get_all_resources(session)
    all_quizzes = await db.get_all_quizzes(session)
    unread_fb = await db.get_unread_feedback_count(session)

    await message.answer(
        texts.ADMIN_STATS.format(
            total=total,
            akademiya=akademiya,
            ilmnur=ilmnur,
            resources=len(all_resources),
            quizzes=len(all_quizzes),
            feedbacks=unread_fb,
        ),
        reply_markup=stats_export_kb(),
        parse_mode="HTML",
    )


@router.callback_query(F.data == "export:csv")
async def export_csv(callback: CallbackQuery, session: AsyncSession):
    """Foydalanuvchilar ro'yxatini CSV formatida eksport qilish."""
    users = await db.get_all_users_for_export(session)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Telegram ID", "Ism", "O'quv markaz", "Faol", "Ro'yxatdan o'tgan"])
    for u in users:
        writer.writerow([
            u.id,
            u.telegram_id,
            u.full_name,
            u.learning_center or "—",
            "Ha" if u.is_active else "Yo'q",
            u.created_at.strftime("%d.%m.%Y %H:%M") if u.created_at else "—",
        ])

    csv_bytes = output.getvalue().encode("utf-8-sig")
    file = BufferedInputFile(
        csv_bytes,
        filename=f"users_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
    )

    await callback.message.answer_document(file, caption="📥 Foydalanuvchilar ro'yxati (CSV)")
    await callback.answer()


@router.callback_query(F.data == "export:xlsx")
async def export_xlsx(callback: CallbackQuery, session: AsyncSession):
    """Foydalanuvchilar ro'yxatini Excel formatida eksport qilish."""
    try:
        from openpyxl import Workbook

        users = await db.get_all_users_for_export(session)

        wb = Workbook()
        ws = wb.active
        ws.title = "Foydalanuvchilar"

        headers = ["ID", "Telegram ID", "Ism", "O'quv markaz", "Faol", "Ro'yxatdan o'tgan"]
        ws.append(headers)

        # Sarlavha stilini qo'yish
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for u in users:
            ws.append([
                u.id,
                u.telegram_id,
                u.full_name,
                u.learning_center or "—",
                "Ha" if u.is_active else "Yo'q",
                u.created_at.strftime("%d.%m.%Y %H:%M") if u.created_at else "—",
            ])

        # Ustun kengligini sozlash
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[col[0].column_letter].width = max_length + 3

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        file = BufferedInputFile(
            buffer.getvalue(),
            filename=f"users_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        )

        await callback.message.answer_document(file, caption="📥 Foydalanuvchilar ro'yxati (Excel)")
        await callback.answer()
    except ImportError:
        await callback.answer("openpyxl o'rnatilmagan!", show_alert=True)
    except Exception:
        await callback.answer("⚠️ Xatolik yuz berdi!", show_alert=True)
