import csv
import io
from datetime import datetime
from typing import Sequence

from aiogram.types import BufferedInputFile

from app.database.models import User


def export_users_csv(users: Sequence[User]) -> BufferedInputFile:
    """Foydalanuvchilar ro'yxatini CSV formatida eksport qilish."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Telegram ID", "Ism", "O'quv markaz", "Faol", "Ro'yxatdan o'tgan"])
    
    for u in users:
        writer.writerow([
            u.id,
            u.telegram_id,
            u.full_name,
            u.learning_center or "—",
            "Ha" if u.is_active else "Yo'q",
            u.created_at.strftime("%d.%m.%Y %H:%M") if u.created_at else "—",
        ])

    csv_bytes = output.getvalue().encode("utf-8-sig")
    return BufferedInputFile(
        csv_bytes,
        filename=f"users_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
    )


def export_users_xlsx(users: Sequence[User]) -> BufferedInputFile:
    """Foydalanuvchilar ro'yxatini Excel formatida eksport qilish."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Foydalanuvchilar"

    headers = ["ID", "Telegram ID", "Ism", "O'quv markaz", "Faol", "Ro'yxatdan o'tgan"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for u in users:
        ws.append([
            u.id,
            u.telegram_id,
            u.full_name,
            u.learning_center or "—",
            "Ha" if u.is_active else "Yo'q",
            u.created_at.strftime("%d.%m.%Y %H:%M") if u.created_at else "—",
        ])

    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col[0].column_letter].width = max_length + 3

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return BufferedInputFile(
        buffer.getvalue(),
        filename=f"users_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
    )
